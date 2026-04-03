from __future__ import annotations

import csv
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from langchain_core.tools import tool
from openpyxl import load_workbook

from app.config import DEFAULT_KNOWLEDGE_GOOGLE_SHEETS_CATALOG_PATH, load_settings
from app.knowledge_paths import build_knowledge_markdown_relative_path
from app.paths import resolve_project_path
from tools.google_workspace_services import get_google_sheets_service


@dataclass(frozen=True)
class KnowledgeSource:
    source_id: str
    source_type: str
    metadata: dict[str, str]
    aliases: tuple[str, ...] = ()
    local_path: Path | None = None
    spreadsheet_id: str = ""
    tabs: tuple[str, ...] = ()
    tab_ranges: dict[str, str] | None = None


@dataclass(frozen=True)
class KnowledgeBlock:
    kind: str
    title: str
    content: str
    start_line: int
    end_line: int


@dataclass(frozen=True)
class KnowledgeDocumentIndex:
    path: str
    metadata: dict[str, str]
    content: str
    lines: list[str]
    blocks: list[KnowledgeBlock]


@dataclass(frozen=True)
class GoogleSheetsKnowledgeCatalogEntry:
    spreadsheet_id: str
    title: str
    aliases: tuple[str, ...]
    tabs: tuple[str, ...]
    tab_ranges: dict[str, str]


@dataclass
class KnowledgeIndexCacheEntry:
    created_at: float
    index: KnowledgeDocumentIndex


_google_sheet_index_cache: dict[str, KnowledgeIndexCacheEntry] = {}


def get_knowledge_base_root() -> Path:
    settings = load_settings()
    return resolve_project_path(settings.knowledge_base_dir)


def get_supported_knowledge_file_types() -> tuple[str, ...]:
    settings = load_settings()
    return tuple(file_type.lower() for file_type in settings.knowledge_file_types)


def get_google_sheets_catalog_path() -> Path:
    settings = load_settings()
    configured_path = getattr(
        settings,
        "knowledge_google_sheets_catalog_path",
        DEFAULT_KNOWLEDGE_GOOGLE_SHEETS_CATALOG_PATH,
    )
    return resolve_project_path(
        configured_path,
        DEFAULT_KNOWLEDGE_GOOGLE_SHEETS_CATALOG_PATH,
    )


def get_google_sheets_cache_ttl_seconds() -> int:
    settings = load_settings()
    return max(int(getattr(settings, "knowledge_google_sheets_cache_ttl_seconds", 120)), 1)


def get_knowledge_document_paths() -> list[Path]:
    root = get_knowledge_base_root()
    if not root.exists() or not root.is_dir():
        return []

    supported_types = set(get_supported_knowledge_file_types())
    documents: list[Path] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if is_hidden_path(path, root):
            continue
        if path.suffix.lower() not in supported_types:
            continue
        documents.append(path)
    return documents


def is_hidden_path(path: Path, root: Path) -> bool:
    try:
        relative_path = path.relative_to(root)
    except ValueError:
        relative_path = path
    return any(part.startswith(".") for part in relative_path.parts)


def get_knowledge_sources() -> list[KnowledgeSource]:
    sources = [build_local_knowledge_source(path) for path in get_knowledge_document_paths()]
    sources.extend(build_google_sheet_knowledge_sources())
    return sources


def build_local_knowledge_source(path: Path) -> KnowledgeSource:
    metadata = build_document_metadata(path)
    metadata["source_type"] = "local_file"
    return KnowledgeSource(
        source_id=f"local:{path.resolve()}",
        source_type="local_file",
        metadata=metadata,
        aliases=(),
        local_path=path,
    )


def build_google_sheet_knowledge_sources() -> list[KnowledgeSource]:
    sources: list[KnowledgeSource] = []
    for entry in load_google_sheets_catalog_entries():
        metadata = {
            "name": entry.title or entry.spreadsheet_id,
            "title": entry.title or entry.spreadsheet_id,
            "path": f"google_sheets/{entry.spreadsheet_id}",
            "file_type": "google_sheet",
            "source_type": "google_sheet",
            "spreadsheet_id": entry.spreadsheet_id,
        }
        sources.append(
            KnowledgeSource(
                source_id=f"google_sheet:{entry.spreadsheet_id}",
                source_type="google_sheet",
                metadata=metadata,
                aliases=entry.aliases,
                spreadsheet_id=entry.spreadsheet_id,
                tabs=entry.tabs,
                tab_ranges=entry.tab_ranges,
            )
        )
    return sources


def load_google_sheets_catalog_entries() -> list[GoogleSheetsKnowledgeCatalogEntry]:
    catalog_path = get_google_sheets_catalog_path()
    if not catalog_path.exists():
        return []

    raw_catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    if isinstance(raw_catalog, dict):
        raw_entries = raw_catalog.get("documents", [])
    elif isinstance(raw_catalog, list):
        raw_entries = raw_catalog
    else:
        raise RuntimeError("Google Sheets knowledge catalog must be a JSON object or list.")

    entries: list[GoogleSheetsKnowledgeCatalogEntry] = []
    for raw_entry in raw_entries:
        if not isinstance(raw_entry, dict):
            continue
        spreadsheet_id = str(raw_entry.get("spreadsheet_id", "")).strip()
        if not spreadsheet_id:
            continue

        title = str(raw_entry.get("title", "")).strip() or spreadsheet_id
        aliases = tuple(
            alias.strip()
            for alias in raw_entry.get("aliases", [])
            if isinstance(alias, str) and alias.strip()
        )
        tabs = tuple(
            tab.strip()
            for tab in raw_entry.get("tabs", [])
            if isinstance(tab, str) and tab.strip()
        )
        raw_ranges = raw_entry.get("ranges", {})
        tab_ranges = {
            str(tab).strip(): str(range_value).strip()
            for tab, range_value in raw_ranges.items()
            if str(tab).strip() and str(range_value).strip()
        } if isinstance(raw_ranges, dict) else {}

        entries.append(
            GoogleSheetsKnowledgeCatalogEntry(
                spreadsheet_id=spreadsheet_id,
                title=title,
                aliases=aliases,
                tabs=tabs,
                tab_ranges=tab_ranges,
            )
        )

    return entries


def load_knowledge_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".txt", ".rst"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix in {".csv", ".tsv"}:
        delimiter = "," if suffix == ".csv" else "\t"
        return render_delimited_file_as_text(path, delimiter=delimiter)
    if suffix in {".xlsx", ".xlsm"}:
        return render_workbook_as_text(path)
    raise ValueError(f"Unsupported knowledge document type: {suffix}")


def build_document_metadata(path: Path) -> dict[str, str]:
    content = load_knowledge_document(path)
    return build_document_metadata_from_content(path, content)


def safe_relative_path(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.name


def humanize_path_name(path: Path) -> str:
    return path.stem.replace("_", " ").replace("-", " ").strip() or path.stem


def extract_document_title(content: str, *, default: str) -> str:
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            return stripped.lstrip("#").strip() or default
        if stripped:
            return stripped
    return default


def build_document_index(source: KnowledgeSource) -> KnowledgeDocumentIndex:
    if source.source_type == "google_sheet":
        return build_google_sheet_document_index(source)
    if source.local_path is None:
        raise RuntimeError(f"Local knowledge source is missing a file path: {source.source_id}")

    content = load_knowledge_document(source.local_path)
    metadata = build_document_metadata_from_content(source.local_path, content)
    metadata["source_type"] = "local_file"
    return build_document_index_from_content(
        document_path=metadata.get("path", source.local_path.name),
        content=content,
        metadata=metadata,
    )


def build_google_sheet_document_index(source: KnowledgeSource) -> KnowledgeDocumentIndex:
    cached_entry = _google_sheet_index_cache.get(source.source_id)
    cache_ttl_seconds = get_google_sheets_cache_ttl_seconds()
    now = time.time()
    if cached_entry is not None and now - cached_entry.created_at < cache_ttl_seconds:
        return cached_entry.index

    content, metadata = fetch_google_sheet_document_content(source)
    index = build_document_index_from_content(
        document_path=metadata.get("path", source.source_id),
        content=content,
        metadata=metadata,
    )
    _google_sheet_index_cache[source.source_id] = KnowledgeIndexCacheEntry(
        created_at=now,
        index=index,
    )
    return index


def build_document_index_from_content(
    *,
    document_path: str,
    content: str,
    metadata: dict[str, str],
) -> KnowledgeDocumentIndex:
    lines = content.splitlines()
    blocks = extract_document_blocks(lines)
    return KnowledgeDocumentIndex(
        path=document_path,
        metadata=metadata,
        content=content,
        lines=lines,
        blocks=blocks,
    )


def build_document_metadata_from_content(path: Path, content: str) -> dict[str, str]:
    root = get_knowledge_base_root()
    relative_path = safe_relative_path(path, root)
    title = extract_document_title(content, default=humanize_path_name(path))
    return {
        "name": path.name,
        "title": title,
        "path": relative_path,
        "file_type": path.suffix.lower(),
    }


def fetch_google_sheet_document_content(source: KnowledgeSource) -> tuple[str, dict[str, str]]:
    if not source.spreadsheet_id:
        raise RuntimeError(f"Google Sheet knowledge source is missing a spreadsheet ID: {source.source_id}")

    service = get_google_sheets_service()
    spreadsheet = service.spreadsheets().get(spreadsheetId=source.spreadsheet_id).execute()
    spreadsheet_properties = spreadsheet.get("properties", {})
    spreadsheet_title = str(spreadsheet_properties.get("title", "")).strip() or source.metadata.get(
        "title",
        source.spreadsheet_id,
    )
    available_tabs = [
        str(sheet.get("properties", {}).get("title", "")).strip()
        for sheet in spreadsheet.get("sheets", [])
        if str(sheet.get("properties", {}).get("title", "")).strip()
    ]
    tabs_to_fetch = resolve_google_sheet_tabs(source, available_tabs)

    if not tabs_to_fetch:
        content = f"# {spreadsheet_title}\n(empty spreadsheet)"
        return content, build_google_sheet_metadata(source, spreadsheet_title)

    requested_ranges = [
        build_google_sheet_a1_range(tab_name, (source.tab_ranges or {}).get(tab_name, ""))
        for tab_name in tabs_to_fetch
    ]
    values_result = (
        service.spreadsheets()
        .values()
        .batchGet(spreadsheetId=source.spreadsheet_id, ranges=requested_ranges)
        .execute()
    )
    value_ranges = values_result.get("valueRanges", [])

    lines = [f"# {spreadsheet_title}"]
    for index, tab_name in enumerate(tabs_to_fetch):
        lines.append(f"## Sheet: {tab_name}")
        value_range = value_ranges[index] if index < len(value_ranges) else {}
        rows = [
            [format_cell_value(cell) for cell in row]
            for row in value_range.get("values", [])
            if isinstance(row, list)
        ]
        lines.extend(render_sheet_rows(rows))

    return "\n".join(line for line in lines if line).strip(), build_google_sheet_metadata(source, spreadsheet_title)


def build_google_sheet_metadata(source: KnowledgeSource, title: str) -> dict[str, str]:
    return {
        "name": title,
        "title": title,
        "path": source.metadata.get("path", f"google_sheets/{source.spreadsheet_id}"),
        "file_type": "google_sheet",
        "source_type": "google_sheet",
        "spreadsheet_id": source.spreadsheet_id,
    }


def resolve_google_sheet_tabs(source: KnowledgeSource, available_tabs: list[str]) -> list[str]:
    if source.tabs:
        requested_tabs = list(source.tabs)
    elif source.tab_ranges:
        requested_tabs = list(source.tab_ranges.keys())
    else:
        requested_tabs = list(available_tabs)

    missing_tabs = [tab for tab in requested_tabs if tab not in available_tabs]
    if missing_tabs:
        raise RuntimeError(
            "Google Sheets knowledge source is configured with missing tabs: "
            + ", ".join(missing_tabs)
        )
    return requested_tabs


def build_google_sheet_a1_range(tab_name: str, configured_range: str) -> str:
    normalized_range = configured_range.strip()
    if not normalized_range:
        return quote_google_sheet_name(tab_name)
    if "!" in normalized_range:
        return normalized_range
    return f"{quote_google_sheet_name(tab_name)}!{normalized_range}"


def quote_google_sheet_name(tab_name: str) -> str:
    escaped_name = tab_name.replace("'", "''")
    return f"'{escaped_name}'"


def extract_document_blocks(lines: list[str]) -> list[KnowledgeBlock]:
    if not lines:
        return []

    heading_entries = [
        (index, heading_level(line), heading_text(line))
        for index, line in enumerate(lines)
        if is_markdown_heading(line)
    ]

    blocks: list[KnowledgeBlock] = []
    if heading_entries:
        start_from = 0
        if heading_entries[0][0] == 0 and heading_entries[0][1] == 1:
            start_from = 1

        for entry_index in range(start_from, len(heading_entries)):
            line_index, level, title = heading_entries[entry_index]
            block_end = len(lines)
            for next_line_index, next_level, _ in heading_entries[entry_index + 1 :]:
                if next_level <= level:
                    block_end = next_line_index
                    break
            block_lines = [line.rstrip() for line in lines[line_index:block_end] if line.strip()]
            if not block_lines:
                continue
            blocks.append(
                KnowledgeBlock(
                    kind=classify_block(block_lines),
                    title=title,
                    content="\n".join(block_lines).strip(),
                    start_line=line_index + 1,
                    end_line=block_end,
                )
            )

    if blocks:
        return blocks

    return extract_freeform_blocks(lines)


def extract_freeform_blocks(lines: list[str]) -> list[KnowledgeBlock]:
    blocks: list[KnowledgeBlock] = []
    line_index = 0
    while line_index < len(lines):
        line = lines[line_index].rstrip()
        if not line.strip():
            line_index += 1
            continue

        if is_markdown_table_start(lines, line_index):
            start = line_index
            table_lines = [line]
            line_index += 1
            while line_index < len(lines) and lines[line_index].strip().startswith("|"):
                table_lines.append(lines[line_index].rstrip())
                line_index += 1
            blocks.append(
                KnowledgeBlock(
                    kind="table",
                    title="Table",
                    content="\n".join(table_lines).strip(),
                    start_line=start + 1,
                    end_line=line_index,
                )
            )
            continue

        start = line_index
        block_lines = [line]
        line_index += 1
        while line_index < len(lines):
            next_line = lines[line_index].rstrip()
            if not next_line.strip():
                break
            if is_markdown_table_start(lines, line_index):
                break
            block_lines.append(next_line)
            line_index += 1

        blocks.append(
            KnowledgeBlock(
                kind=classify_block(block_lines),
                title=block_title_from_lines(block_lines),
                content="\n".join(block_lines).strip(),
                start_line=start + 1,
                end_line=line_index,
            )
        )

    return blocks


def heading_text(line: str) -> str:
    return re.sub(r"^\s*#{1,6}\s+", "", line).strip()


def classify_block(lines: list[str]) -> str:
    body_lines = [line.strip() for line in lines if line.strip()]
    if not body_lines:
        return "paragraph"
    if any(line.startswith("|") for line in body_lines):
        return "table"
    if any(is_list_like_line(line) for line in body_lines[1:] or body_lines):
        return "list"
    return "section"


def block_title_from_lines(lines: list[str]) -> str:
    for line in lines:
        stripped = line.strip()
        if stripped:
            return stripped[:80]
    return "Block"


def is_list_like_line(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("- ") or is_numbered_item(stripped) or is_bullet_item(stripped)


def is_markdown_table_start(lines: list[str], index: int) -> bool:
    if index + 1 >= len(lines):
        return False
    current = lines[index].strip()
    next_line = lines[index + 1].strip()
    return current.startswith("|") and bool(re.match(r"^\|\s*:?-{3,}", next_line))


def score_block(block: KnowledgeBlock, query: str, terms: list[str]) -> int:
    combined_text = "\n".join(part for part in (block.title, block.content) if part).lower()
    score = 0
    normalized_query = query.strip().lower()
    if normalized_query and normalized_query in combined_text:
        score += max(8, len(terms) * 2)

    for term in terms:
        score += combined_text.count(term)

    if score > 0 and block.kind == "section":
        score += 1
    return score


def find_best_matching_block(
    blocks: list[KnowledgeBlock],
    query: str,
    terms: list[str],
    *,
    preferred_kinds: tuple[str, ...] = (),
) -> KnowledgeBlock | None:
    best_block: KnowledgeBlock | None = None
    best_score = 0
    best_span = 0

    for block in blocks:
        score = score_block(block, query, terms)
        if preferred_kinds and block.kind in preferred_kinds:
            score += 2
        span = max(block.end_line - block.start_line + 1, 1)
        if score > best_score or (score == best_score and score > 0 and (best_block is None or span < best_span)):
            best_score = score
            best_block = block
            best_span = span

    return best_block if best_score > 0 else None


def build_block_snippet(block: KnowledgeBlock, *, max_lines: int = 5) -> str:
    lines = [line.rstrip() for line in block.content.splitlines() if line.strip()]
    return "\n".join(lines[:max_lines]).strip()


def build_document_excerpt(
    *,
    document_index: KnowledgeDocumentIndex,
    section_query: str,
    max_lines: int,
) -> dict[str, object]:
    normalized_query = section_query.strip()
    if normalized_query:
        terms = normalize_query_terms(normalized_query)
        preferred_block = find_best_matching_block(
            document_index.blocks,
            normalized_query,
            terms,
            preferred_kinds=("section", "list", "table"),
        )
        if preferred_block is not None:
            excerpt_lines = preferred_block.content.splitlines()[:max_lines]
            return {
                "start_line": preferred_block.start_line,
                "end_line": preferred_block.start_line + len(excerpt_lines) - 1,
                "content": "\n".join(excerpt_lines).strip(),
                "truncated": len(preferred_block.content.splitlines()) > max_lines,
                "block_type": preferred_block.kind,
                "section_title": preferred_block.title,
            }

    if document_index.blocks:
        excerpt_lines: list[str] = []
        start_line = document_index.blocks[0].start_line
        end_line = start_line
        consumed_blocks = 0
        for block in document_index.blocks:
            block_lines = [line.rstrip() for line in block.content.splitlines() if line.strip()]
            remaining = max_lines - len(excerpt_lines)
            if remaining <= 0:
                break
            excerpt_lines.extend(block_lines[:remaining])
            end_line = block.start_line + min(len(block_lines), remaining) - 1
            consumed_blocks += 1
            if len(excerpt_lines) >= max_lines:
                break
        total_lines = sum(
            len([line for line in block.content.splitlines() if line.strip()])
            for block in document_index.blocks
        )
        return {
            "start_line": start_line,
            "end_line": end_line,
            "content": "\n".join(excerpt_lines).strip(),
            "truncated": total_lines > len(excerpt_lines),
            "block_type": document_index.blocks[0].kind,
            "section_title": document_index.blocks[0].title if consumed_blocks == 1 else "",
        }

    lines = document_index.lines
    excerpt_lines = [line.rstrip() for line in lines[:max_lines] if line.strip()]
    return {
        "start_line": 1,
        "end_line": len(excerpt_lines),
        "content": "\n".join(excerpt_lines).strip(),
        "truncated": len(lines) > len(excerpt_lines),
        "block_type": "document",
        "section_title": document_index.metadata.get("title", ""),
    }


def normalize_query_terms(query: str) -> list[str]:
    return [term for term in re.findall(r"[\w\u4e00-\u9fff]+", query.lower()) if len(term) >= 2]


def score_document(content: str, query: str, terms: list[str]) -> int:
    lowered = content.lower()
    score = 0
    normalized_query = query.strip().lower()
    if normalized_query and normalized_query in lowered:
        score += max(8, len(terms) * 2)

    for term in terms:
        score += lowered.count(term)

    return score


def find_best_match_line(lines: list[str], query: str, terms: list[str]) -> int:
    normalized_query = query.strip().lower()
    best_index = 0
    best_score = -1

    for index, line in enumerate(lines):
        lowered = line.lower()
        line_score = 0
        if normalized_query and normalized_query in lowered:
            line_score += max(8, len(terms) * 2)
        for term in terms:
            line_score += lowered.count(term)
        if line_score > best_score:
            best_score = line_score
            best_index = index

    return best_index


def build_snippet(lines: list[str], match_index: int, *, radius: int = 2) -> str:
    start = max(match_index - radius, 0)
    end = min(match_index + radius + 1, len(lines))
    snippet_lines = [line.rstrip() for line in lines[start:end] if line.strip()]
    return "\n".join(snippet_lines).strip()


def is_markdown_heading(line: str) -> bool:
    return bool(re.match(r"^\s*#{1,6}\s+", line))


def heading_level(line: str) -> int:
    match = re.match(r"^\s*(#{1,6})\s+", line)
    if not match:
        return 0
    return len(match.group(1))


def resolve_document(document_name: str) -> KnowledgeSource | None:
    normalized_query = document_name.strip().lower()
    if not normalized_query:
        return None

    exact_matches: list[KnowledgeSource] = []
    partial_matches: list[KnowledgeSource] = []

    for source in get_knowledge_sources():
        candidates = {
            str(value).strip().lower()
            for value in [
                source.metadata.get("name", ""),
                source.metadata.get("title", ""),
                source.metadata.get("path", ""),
                source.spreadsheet_id,
                *(alias for alias in source.aliases),
            ]
            if str(value).strip()
        }
        if source.local_path is not None:
            candidates.add(source.local_path.name.lower())
            candidates.add(source.local_path.stem.lower())

        if normalized_query in candidates:
            exact_matches.append(source)
            continue
        if any(normalized_query in candidate for candidate in candidates):
            partial_matches.append(source)

    if exact_matches:
        return exact_matches[0]
    if partial_matches:
        return partial_matches[0]
    return None


def locate_section(lines: list[str], section_query: str) -> tuple[int, int]:
    normalized_query = section_query.strip().lower()
    if not normalized_query:
        return 0, len(lines)

    for index, line in enumerate(lines):
        if normalized_query not in line.lower():
            continue

        if is_markdown_heading(line):
            section_start = index
            section_level = heading_level(line)
            section_end = len(lines)
            for next_index in range(index + 1, len(lines)):
                next_line = lines[next_index]
                if is_markdown_heading(next_line) and heading_level(next_line) <= section_level:
                    section_end = next_index
                    break
            return section_start, section_end

        return max(index - 3, 0), min(index + 4, len(lines))

    return 0, len(lines)


def render_delimited_file_as_text(path: Path, *, delimiter: str) -> str:
    title = humanize_path_name(path)
    lines = [f"# {title}"]
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = [[cell.strip() for cell in row] for row in reader]
        lines.extend(render_sheet_rows(rows))
    return "\n".join(line for line in lines if line).strip()


def render_workbook_as_text(path: Path) -> str:
    title = humanize_path_name(path)
    lines = [f"# {title}"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            lines.append("(empty workbook)")
            return "\n".join(lines)

        for worksheet in workbook.worksheets:
            lines.append(f"## Sheet: {worksheet.title}")
            rows = [
                [format_cell_value(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            lines.extend(render_sheet_rows(rows))
    finally:
        workbook.close()

    return "\n".join(line for line in lines if line).strip()


def render_sheet_rows(rows: Iterable[list[str]]) -> list[str]:
    normalized_rows = [trim_trailing_empty_cells(list(row)) for row in rows]
    rendered_lines: list[str] = []
    row_index = 0

    while row_index < len(normalized_rows):
        row = normalized_rows[row_index]
        if not has_non_empty_cells(row):
            row_index += 1
            continue

        if is_table_block_start(normalized_rows, row_index):
            block_rows, row_index = consume_table_block(normalized_rows, row_index)
            rendered_lines.extend(render_table_block(block_rows))
            continue

        rendered_line = render_sparse_row(row)
        if rendered_line:
            rendered_lines.append(rendered_line)
        row_index += 1

    if not rendered_lines:
        rendered_lines.append("(empty)")

    return rendered_lines


def trim_trailing_empty_cells(row: list[str]) -> list[str]:
    trimmed = list(row)
    while trimmed and not trimmed[-1].strip():
        trimmed.pop()
    return trimmed


def has_non_empty_cells(row: list[str]) -> bool:
    return any(cell.strip() for cell in row)


def get_non_empty_positions(row: list[str]) -> list[int]:
    return [index for index, cell in enumerate(row) if cell.strip()]


def render_sparse_row(row: list[str]) -> str:
    positions = get_non_empty_positions(row)
    if not positions:
        return ""

    if len(positions) == 1:
        return render_single_value_row(row[positions[0]].strip(), positions[0])

    indent = "  " * max(positions[0] - 1, 0)
    values = [cell.strip() for cell in row if cell.strip()]
    return f"{indent}{' | '.join(escape_table_cell(value) for value in values)}"


def render_single_value_row(value: str, column_index: int) -> str:
    indent = "  " * max(column_index - 1, 0)
    cleaned = value.strip()
    if not cleaned:
        return ""

    if is_bullet_item(cleaned):
        return f"{indent}- {strip_bullet_marker(cleaned)}"
    if is_numbered_item(cleaned):
        return f"{indent}{cleaned}"

    heading_level = min(6, 2 + max(column_index - 1, 0))
    return f"{'#' * heading_level} {cleaned.rstrip('：:').strip()}"


def is_bullet_item(value: str) -> bool:
    return bool(re.match(r"^[·•●▪◦\-*+]\s*", value))


def strip_bullet_marker(value: str) -> str:
    return re.sub(r"^[·•●▪◦\-*+]\s*", "", value).strip()


def is_numbered_item(value: str) -> bool:
    return bool(
        re.match(r"^(?:[（(]?\d+[)）.、]|[A-Za-z]\.)\s*", value)
    )


def is_table_block_start(rows: list[list[str]], start_index: int) -> bool:
    current_positions = get_non_empty_positions(rows[start_index])
    if len(current_positions) <= 1:
        return False
    if start_index + 1 >= len(rows):
        return False
    next_positions = get_non_empty_positions(rows[start_index + 1])
    return len(next_positions) > 1


def consume_table_block(rows: list[list[str]], start_index: int) -> tuple[list[list[str]], int]:
    block: list[list[str]] = []
    row_index = start_index
    while row_index < len(rows):
        row = rows[row_index]
        if not has_non_empty_cells(row):
            break
        if len(get_non_empty_positions(row)) <= 1:
            break
        block.append(row)
        row_index += 1
    return block, row_index


def render_table_block(rows: list[list[str]]) -> list[str]:
    min_column = min(get_non_empty_positions(row)[0] for row in rows)
    max_column = max(get_non_empty_positions(row)[-1] for row in rows)

    trimmed_rows = [
        [row[index].strip() if index < len(row) else "" for index in range(min_column, max_column + 1)]
        for row in rows
    ]
    header = build_table_header(trimmed_rows[0])
    rendered_lines = [
        f"| {' | '.join(header)} |",
        f"| {' | '.join('---' for _ in header)} |",
    ]

    previous_first_value = ""
    for row in trimmed_rows[1:]:
        padded_row = row + [""] * (len(header) - len(row))
        if not any(cell.strip() for cell in padded_row):
            continue
        if not padded_row[0].strip() and previous_first_value:
            padded_row[0] = previous_first_value
        if padded_row[0].strip():
            previous_first_value = padded_row[0].strip()
        rendered_lines.append(
            f"| {' | '.join(escape_table_cell(cell.strip()) or ' ' for cell in padded_row[:len(header)])} |"
        )

    return rendered_lines


def build_table_header(row: list[str]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        header_value = value.strip() or f"column_{index}"
        headers.append(escape_table_cell(header_value))
    return headers


def escape_table_cell(value: str) -> str:
    return value.replace("|", "/").strip()


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def build_knowledge_base_context() -> dict[str, object]:
    root = get_knowledge_base_root()
    catalog_path = get_google_sheets_catalog_path()
    return {
        "knowledge_base_dir": str(root),
        "exists": root.exists() and root.is_dir(),
        "supported_file_types": list(get_supported_knowledge_file_types()),
        "google_sheets_catalog_path": str(catalog_path),
        "google_sheets_catalog_exists": catalog_path.exists(),
    }


def resolve_knowledge_markdown_target(relative_path: str) -> tuple[Path, Path]:
    root = get_knowledge_base_root().resolve()
    candidate = (root / str(relative_path or "").strip()).resolve()
    candidate.relative_to(root)
    if candidate.suffix.lower() != ".md":
        raise ValueError("Only Markdown files ending in .md can be written by this tool.")
    return root, candidate


@tool
def list_knowledge_documents() -> dict[str, object]:
    """List knowledge documents available to the knowledge agent."""
    context = build_knowledge_base_context()
    try:
        documents = [source.metadata for source in get_knowledge_sources()]
    except Exception as exc:
        return {
            "ok": False,
            **context,
            "error": str(exc),
            "document_count": 0,
            "documents": [],
        }

    return {
        "ok": True,
        **context,
        "document_count": len(documents),
        "documents": documents,
    }


@tool
def search_knowledge_documents(query: str, limit: int = 5) -> dict[str, object]:
    """Search knowledge documents from local files and curated online sheets."""
    normalized_limit = max(limit, 1)
    context = build_knowledge_base_context()
    if not query.strip():
        return {
            "ok": False,
            **context,
            "error": "Query cannot be empty.",
            "query": query,
            "match_count": 0,
            "documents": [],
        }

    terms = normalize_query_terms(query)
    matches: list[dict[str, object]] = []

    try:
        sources = get_knowledge_sources()
    except Exception as exc:
        return {
            "ok": False,
            **context,
            "error": str(exc),
            "query": query,
            "match_count": 0,
            "documents": [],
        }

    for source in sources:
        document_index = build_document_index(source)
        best_block = find_best_matching_block(document_index.blocks, query, terms)
        if best_block is None:
            continue
        score = score_block(best_block, query, terms)
        matches.append(
            {
                **document_index.metadata,
                "score": score,
                "line_number": best_block.start_line,
                "end_line": best_block.end_line,
                "snippet": build_block_snippet(best_block),
                "block_type": best_block.kind,
                "section_title": best_block.title,
            }
        )

    matches.sort(key=lambda item: (-int(item["score"]), str(item["path"])))
    return {
        "ok": True,
        **context,
        "query": query,
        "match_count": len(matches),
        "documents": matches[:normalized_limit],
    }


@tool
def read_knowledge_document(document_name: str, section_query: str = "", max_lines: int = 80) -> dict[str, object]:
    """Read a knowledge document, optionally focusing on a matching section."""
    context = build_knowledge_base_context()
    try:
        source = resolve_document(document_name)
    except Exception as exc:
        return {
            "ok": False,
            **context,
            "error": str(exc),
            "document_name": document_name,
        }

    if source is None:
        return {
            "ok": False,
            **context,
            "error": f"Document not found: {document_name}",
            "document_name": document_name,
        }

    normalized_max_lines = max(max_lines, 1)
    document_index = build_document_index(source)
    excerpt = build_document_excerpt(
        document_index=document_index,
        section_query=section_query,
        max_lines=normalized_max_lines,
    )

    return {
        "ok": True,
        **context,
        "document": document_index.metadata,
        "section_query": section_query.strip(),
        "start_line": excerpt["start_line"],
        "end_line": excerpt["end_line"],
        "content": excerpt["content"],
        "truncated": excerpt["truncated"],
        "block_type": excerpt["block_type"],
        "section_title": excerpt["section_title"],
    }


@tool
def resolve_knowledge_markdown_path(
    layer: str,
    category: str,
    filename: str = "README.md",
    game_slug: str = "",
    market_slug: str = "",
    feature_slug: str = "",
    legacy_bucket: str = "",
) -> dict[str, object]:
    """Resolve a canonical Markdown path inside the knowledge base hierarchy."""
    context = build_knowledge_base_context()
    try:
        relative_path = build_knowledge_markdown_relative_path(
            layer=layer,
            category=category,
            filename=filename,
            game_slug=game_slug,
            market_slug=market_slug,
            feature_slug=feature_slug,
            legacy_bucket=legacy_bucket,
        )
        root, absolute_path = resolve_knowledge_markdown_target(relative_path)
    except Exception as exc:
        return {
            "ok": False,
            **context,
            "error": str(exc),
            "layer": layer,
            "category": category,
        }

    return {
        "ok": True,
        **context,
        "layer": layer,
        "category": category,
        "relative_path": relative_path,
        "absolute_path": str(absolute_path),
        "knowledge_root": str(root),
    }


@tool
def write_knowledge_markdown_document(
    relative_path: str,
    content: str,
    overwrite: bool = False,
) -> dict[str, object]:
    """Write a Markdown document inside the knowledge base root."""
    context = build_knowledge_base_context()
    if not str(content or "").strip():
        return {
            "ok": False,
            **context,
            "error": "Document content cannot be empty.",
            "relative_path": relative_path,
        }

    try:
        _root, absolute_path = resolve_knowledge_markdown_target(relative_path)
    except Exception as exc:
        return {
            "ok": False,
            **context,
            "error": str(exc),
            "relative_path": relative_path,
        }

    existed = absolute_path.exists()
    if existed and not overwrite:
        return {
            "ok": False,
            **context,
            "error": "Target file already exists. Retry with overwrite=True to replace it.",
            "relative_path": relative_path,
            "absolute_path": str(absolute_path),
        }

    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    absolute_path.write_text(content, encoding="utf-8")

    return {
        "ok": True,
        **context,
        "relative_path": relative_path,
        "absolute_path": str(absolute_path),
        "created": not existed,
        "overwritten": existed,
        "bytes_written": absolute_path.stat().st_size,
    }
